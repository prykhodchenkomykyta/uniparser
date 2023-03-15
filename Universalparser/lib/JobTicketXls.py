######################################################## Start of Python Script#################################################################################
# Import the openpyxl library
import json
import openpyxl
import sys
import shutil

spython_sJobTicketConfigFile = sys.argv[0]
sxlsx_MasterPID = sys.argv[1]
sxlsx_SchemaJobTicket = sys.argv[2]
sxlsx_OutputJobTicketFile = sys.argv[3]
aCellIndex_FieldsToWrite_MasterPID = sys.argv[4]
aFieldsToWrite_MasterPID = sys.argv[5]
aCellIndex_FieldsToWrite_ProductConfig = sys.argv[6]
aFieldsToWrite_ProductConfig = sys.argv[7]
aMappingFieldNames_ProdConfig_MasterPID = sys.argv[8]
aMappingFieldValues_ProdConfig_MasterPID = sys.argv[9]
sSheetName_MasterPID = sys.argv[10]

aCellIndex_FieldsToWrite_MasterPID = json.loads(aCellIndex_FieldsToWrite_MasterPID)
aFieldsToWrite_MasterPID = json.loads(aFieldsToWrite_MasterPID)
aCellIndex_FieldsToWrite_ProductConfig = json.loads(aCellIndex_FieldsToWrite_ProductConfig)
aFieldsToWrite_ProductConfig = json.loads(aFieldsToWrite_ProductConfig)
aMappingFieldNames_ProdConfig_MasterPID = json.loads(aMappingFieldNames_ProdConfig_MasterPID)
aMappingFieldValues_ProdConfig_MasterPID = json.loads(aMappingFieldValues_ProdConfig_MasterPID)

#Create a clone of schema in the temp folder
shutil.copy(sxlsx_SchemaJobTicket, sxlsx_OutputJobTicketFile)

#Load MasterPID List
xlsx_MasterPID = openpyxl.load_workbook(sxlsx_MasterPID)

#Select the sheet of MasterPID
sheet = xlsx_MasterPID[sSheetName_MasterPID]

#Find column location of all mapping fields in MasterPID
TotalMappedElements = 0
for element in aMappingFieldNames_ProdConfig_MasterPID:
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for c in col:
            if c.value == element:
                aMappingFieldNames_ProdConfig_MasterPID[aMappingFieldNames_ProdConfig_MasterPID.index(element)] = c.column
                break

# Loop over the rows (max: 50 cells in a row) of the sheet and search for the values of mapping field's values
TotalMappedElements = len(aMappingFieldNames_ProdConfig_MasterPID)
MasterPID_Row = None
index = 0
for row_number, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=50,values_only=True), start=2):
    counter = 0    
    for i in range(index, TotalMappedElements):
        test =str(row[aMappingFieldNames_ProdConfig_MasterPID[i]-1]).lower()
        test2 = (aMappingFieldValues_ProdConfig_MasterPID[i]).lower()
        if str(row[aMappingFieldNames_ProdConfig_MasterPID[i]-1]).lower() == aMappingFieldValues_ProdConfig_MasterPID[i].lower():
            counter += 1
        else:
            break
    if counter == TotalMappedElements:
        MasterPID_Row = row_number
        break

if MasterPID_Row == None:
    print('Product code not be found in Master Product List.')

# Prepare data in array:DataToWrite_MasterPID from MasterPID to be written in Job-Ticket
DataToWrite_MasterPID = []
column = str(MasterPID_Row)
for element in aFieldsToWrite_MasterPID:
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for c in col:
            if c.value == element:
                DataToWrite_MasterPID.append(sheet[c.column_letter+column].value)
                break
        else:
            continue
        break

#Load Job-Ticket
JobTicketFile = openpyxl.load_workbook(sxlsx_OutputJobTicketFile)
worksheet = JobTicketFile.active

#Update Job-Ticket with data from Product-Config file
for index, element in enumerate(aFieldsToWrite_ProductConfig):
    worksheet[aCellIndex_FieldsToWrite_ProductConfig[index]] = element

#Update Job-Ticket with data from MasterPID
for index, element in enumerate(DataToWrite_MasterPID):
    worksheet[aCellIndex_FieldsToWrite_MasterPID[index]] = element

#Save the Job-Ticket in output location
JobTicketFile.save(sxlsx_OutputJobTicketFile)

print('TRUE')